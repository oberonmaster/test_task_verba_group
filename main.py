from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook
import re


target_url = "https://www.wildberries.ru/catalog/0/search.aspx?search=пальто%20из%20натуральной%20шерсти"

# Настройки selenium
chrome_options = Options()
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
driver = webdriver.Chrome(options=chrome_options)
driver.get(target_url)


wait = WebDriverWait(driver, 60)
wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-card")))


# Даем время на полную загрузку
time.sleep(3)


html = driver.page_source
cards = driver.find_elements(By.CSS_SELECTOR, ".product-card")


products = {}


# переход по карточкам товаров
for card in cards:
    # переход на карточку текуoего товара
    link_element = card.find_element(By.CSS_SELECTOR, ".product-card__link.j-card-link")
    
    product_url = link_element.get_attribute("href")
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[1])
    driver.get(product_url)

    time.sleep(5)    
    
    # информация с карточки товара
    article = driver.current_url.split("/")[-2]
    product_name = driver.find_elements(By.CSS_SELECTOR, "[class*='productTitle']")[0].text
    product_price = driver.find_elements(By.CSS_SELECTOR, "[class*='priceBlockFinalPrice']")[0].text
    product_rating = driver.find_element(By.XPATH, '//*[@id="reactContainers"]/div[2]/div/div[2]/div[2]/div[3]/div[1]/a[1]/div[1]/div/p[1]').text
    product_feedback_count = re.sub(r'[^\d.-]', '', driver.find_element(By.XPATH, '//*[@id="reactContainers"]/div[2]/div/div[2]/div[2]/div[3]/div[1]/a[1]/div[1]/div/p[2]').text)
    
    seller = driver.find_element(By.CSS_SELECTOR, "h4[class*='sellerAndBrandItemName']").text
    # seller_link = driver.find_element(By.XPATH, '//*[@id="reactContainers"]/div[2]/div/div[2]/div[2]/div[4]/div/div[5]/div/div/a').get_attribute("href")
    # Ссылка на продавца
    try:
        # Ищем ссылку разными способами
        seller_link_elem = driver.find_element(By.CSS_SELECTOR, "a[href*='/seller/'], a[href*='/brand/'], a.seller-info__link")
        seller_link = seller_link_elem.get_attribute("href")
    except:
        try:
            # Альтернативный поиск
            seller_link_elem = driver.find_element(By.XPATH, "//a[contains(@href, '/seller/') or contains(@href, '/brand/')]")
            seller_link = seller_link_elem.get_attribute("href")
        except:
            seller_link = ""

    
    # Все размеры
    size_elements = driver.find_elements(By.CSS_SELECTOR, "span[class*='sizesListSize'], span[class*='sizesListSizeRu']")
    all_numbers = set()
    for element in size_elements:
        text = element.text.strip()
        numbers = re.findall(r'\d+', text)
        all_numbers.update(numbers)
        
    if all_numbers:
        sorted_sizes = sorted(all_numbers, key=int)
        product_sizes = ", ".join(sorted_sizes)
    else:
        product_sizes = ""
        
    # количество
    product_count = ""
    
    # переход по "о товаре"
    button_about = driver.find_element(By.CSS_SELECTOR, "[class*='moreAboutButton']")
    driver.execute_script("arguments[0].click();", button_about)
    time.sleep(5)
    
    # информация с "о товаре"
    # short_description = f"{driver.find_elements(By.CSS_SELECTOR, "[class*='descriptionText']")[0].text[0:20]}..."
    try:
        short_description = driver.find_element(By.CSS_SELECTOR, ".product-page__text, [class*='descriptionText'], .details__content").text[:100] + "..."
    except:
        short_description = ""
    # full_description = driver.find_elements(By.CSS_SELECTOR, "[class*='descriptionText']")[0].text
    try:
        full_description = driver.find_element(By.CSS_SELECTOR, ".product-page__text, [class*='descriptionText'], .details__content").text
    except:
        full_description = ""
    
    features_0 = driver.find_element(By.XPATH, '//*[@id="characteristics"]/table[2]').text
    features_1 = driver.find_element(By.XPATH, '//*[@id="characteristics"]/table[3]').text
    all_features = f"{features_0}/n{features_1}"
    
    # выход из "о товаре"
    driver.switch_to.active_element.send_keys(Keys.ESCAPE)
    time.sleep(5)

    # images_links
    image = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "img[alt='Product image 1']"))
    )
    image.click()
    
    images = driver.find_elements(By.CSS_SELECTOR, ".miniaturesWrapper--Yw0YN img")
    image_urls = [img.get_attribute("src") for img in images if img.get_attribute("src")]
    images_links = ",".join(image_urls)
    driver.switch_to.active_element.send_keys(Keys.ESCAPE)
    time.sleep(5)
    
    
    product_info = {
        "Ссылка на товар" : product_url,
        "Артикул" : article,
        "Название" : product_name,
        "Цена" : product_price,
        "Описание" : short_description,
        "Ссылки на изображения через запятую" : images_links,
        "Описание полное" : full_description,
        "Все характеристики" : all_features,
        "Название селлера" : seller,
        "Ссылка на селлера" : seller_link,
        "Размеры товара через запятую" : product_sizes,
        "Остатки по товару (число)" : product_count,
        "Рейтинг" : product_rating,
        "Количество отзывов" : product_feedback_count,
    }
    
    products[article] = product_info
    driver.close()
    driver.switch_to.window(driver.window_handles[0])


wb = Workbook()
ws = wb.active

# Заголовки
headers = ["Ссылка на товар", 
           "Артикул",
           "Название",
           "Цена",
           "Описание",
           "Ссылки на изображения через запятую",
           "Описание полное",
           "Все характеристики",
           "Название селлера",
           "Ссылка на селлера",
           "Размеры товара через запятую",
           "Остатки по товару (число)",
           "Рейтинг",
           "Количество отзывов"]


for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Данные
row = 2
for article, info in products.items():
    ws.cell(row=row, column=1, value=info.get("Ссылка на товар"))
    ws.cell(row=row, column=2, value=info.get("Артикул"))
    ws.cell(row=row, column=3, value=info.get("Название"))
    ws.cell(row=row, column=4, value=info.get("Цена"))
    ws.cell(row=row, column=5, value=info.get("Описание"))
    ws.cell(row=row, column=6, value=info.get("Ссылки на изображения через запятую"))
    ws.cell(row=row, column=7, value=info.get("Описание полное"))
    ws.cell(row=row, column=8, value=info.get("Все характеристики"))
    ws.cell(row=row, column=9, value=info.get("Название селлера"))
    ws.cell(row=row, column=10, value=info.get("Ссылка на селлера"))
    ws.cell(row=row, column=11, value=info.get("Размеры товара через запятую"))
    ws.cell(row=row, column=12, value=info.get("Остатки по товару (число)"))
    ws.cell(row=row, column=13, value=info.get("Рейтинг"))
    ws.cell(row=row, column=14, value=info.get("Количество отзывов"))
    row += 1

# Сохраняем
wb.save("result.xlsx")
  

driver.quit()

